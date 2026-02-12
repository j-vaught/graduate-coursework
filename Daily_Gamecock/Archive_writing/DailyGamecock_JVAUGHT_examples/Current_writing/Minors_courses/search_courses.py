import openpyxl
import os
from pathlib import Path
from collections import defaultdict

# Courses to search for
TARGET_COURSES = ['571', '567', '566']  # MUSC course numbers
TARGET_SUBJECT = 'MUSC'

# Directory with grade spreads
GRADE_SPREAD_DIR = Path(__file__).parent / 'usc_grade_spreads'

# Results storage
results = defaultdict(list)

print(f"Searching for {TARGET_SUBJECT} {', '.join(TARGET_COURSES)} courses...")
print(f"Scanning: {GRADE_SPREAD_DIR}\n")

# Get all Excel files
excel_files = sorted(GRADE_SPREAD_DIR.glob('*_grade_spread_report*.xlsx'))
print(f"Found {len(excel_files)} grade spread files\n")

# Search each file
for excel_file in excel_files:
    # Extract semester from filename (e.g., 202505 from 202505_grade_spread_report.xlsx)
    filename = excel_file.name
    semester_code = filename.split('_')[0]
    
    # Parse semester code (YYYYMM format)
    year = semester_code[:4]
    month = semester_code[4:6]
    
    semester_map = {
        '01': 'Spring', '05': 'Summer', '08': 'Fall'
    }
    semester_name = semester_map.get(month, f'Month {month}')
    semester_display = f"{semester_name} {year}"
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Read header row to find column indices
        headers = {}
        for col_num, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_num
        
        # Search for target courses
        for row in ws.iter_rows(min_row=2, values_only=False):
            subject = row[headers.get('SUBJECT', 1) - 1].value
            course_num = row[headers.get('COURSE_NUMBER', 2) - 1].value
            section = row[headers.get('COURSE_SECTION_NUMBER', 3) - 1].value
            title = row[headers.get('TITLE', 4) - 1].value
            num_grades = row[headers.get('Num Grades Posted', 34) - 1].value
            
            if subject == TARGET_SUBJECT and str(course_num) in TARGET_COURSES:
                results[f"{TARGET_SUBJECT} {course_num}"].append({
                    'semester': semester_display,
                    'section': section,
                    'title': title,
                    'enrollment': num_grades if num_grades else 0,
                    'file': filename
                })
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading {filename}: {e}")

# Display results
print("="*80)
print("COURSE SEARCH RESULTS")
print("="*80)

if not any(results.values()):
    print("\n❌ NO OFFERINGS FOUND for any of the target courses!\n")
else:
    for course_code in sorted(TARGET_COURSES):
        full_code = f"{TARGET_SUBJECT} {course_code}"
        print(f"\n📚 {full_code}")
        print("-" * 80)
        
        if full_code not in results or not results[full_code]:
            print(f"   ❌ NEVER OFFERED (0 semesters)")
        else:
            offerings = results[full_code]
            print(f"   ✓ Offered {len(offerings)} times across {len(set(o['semester'] for o in offerings))} different semesters\n")
            
            # Group by semester
            by_semester = defaultdict(list)
            for offering in offerings:
                by_semester[offering['semester']].append(offering)
            
            for semester in sorted(by_semester.keys(), key=lambda x: (x.split()[-1], x.split()[0])):
                offerings_in_sem = by_semester[semester]
                total_enrollment = sum(o['enrollment'] for o in offerings_in_sem)
                print(f"   {semester}:")
                for offering in offerings_in_sem:
                    print(f"      Section {offering['section']}: {offering['title']} ({offering['enrollment']} students)")
                print(f"      Total enrollment that semester: {total_enrollment}")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

total_offerings = sum(len(v) for v in results.values())
total_courses_with_offerings = len([v for v in results.values() if v])
print(f"\nTotal courses searched: 3")
print(f"Courses with any offerings: {total_courses_with_offerings}")
print(f"Total course offerings (sections): {total_offerings}")
print(f"Total semesters analyzed: {len(excel_files)}")

if total_courses_with_offerings < 3:
    missing_courses = [f"{TARGET_SUBJECT} {c}" for c in TARGET_COURSES 
                      if f"{TARGET_SUBJECT} {c}" not in results or not results[f"{TARGET_SUBJECT} {c}"]]
    print(f"\n⚠️  COURSES NEVER OFFERED: {', '.join(missing_courses)}")
    print(f"\nThese courses are REQUIRED for Audio Recording Minor but NEVER appear in grade spread data.")
